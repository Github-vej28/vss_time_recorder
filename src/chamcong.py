import datetime
import pandas as pd
import xlrd
from xlutils.copy import copy
import openpyxl
import os
import calendar
from gui import GUI

from employee import Employee

class Timekeeper():
    def __init__(self):
        self.screen = GUI()
        self.file_input = None
        self.file_on_leave = None
        self.file_output = None

        self.employees = dict()
        self.sat = None
        self.sun = None

        self.MONTH = None

    def getWeekendDays(self, start, end):
        saturdays = pd.date_range(start = start, end = end, freq = 'W-SAT')
        sundays = pd.date_range(start = start, end = end, freq = 'W-SUN')
        return saturdays, sundays

    def skipBlankRows(self, sheet):
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=row, column=col).value != None:
                    return row, col

    def getExcelFiles(self):
        self.screen.show()
        self.file_input, self.file_on_leave, self.file_output = self.screen.getFiles()

    def read_timekeeping_machine(self):
        workbook_in = openpyxl.load_workbook(self.file_input)
        sheet = workbook_in.active
        max = sheet.max_row + 1
        start_row, start_col = self.skipBlankRows(sheet)

        tmp = sheet.cell(row = start_row, column = start_col + 1).value
        self.MONTH = tmp.month
        first_day_of_month = tmp.replace(day=1).date()
        last_day_of_month = str(self.MONTH) + "/" + str(calendar.monthrange(tmp.year, self.MONTH)[1]) +"/" + str(tmp.year)
        saturdays, sundays = self.getWeekendDays(first_day_of_month, last_day_of_month)
        self.sat = saturdays.day
        self.sun = sundays.day

        for i in range(start_row, max):
            id = sheet.cell(row = i, column = start_col).value
            if id not in self.employees:
                new_employee = Employee(id)
                self.employees[id] = new_employee
            
            date = sheet.cell(row = i, column = start_col + 1).value
            employee = self.employees[id]
            employee.createDate(date)

            timestamp = sheet.cell(row = i, column = start_col + 2).value
            employee.Timekeeping[date.day].updateTime(timestamp)

    def readFileOnLeave(self):
        work_book_off = openpyxl.load_workbook(self.file_on_leave)
        sheet_off = work_book_off.active
        num_rows = sheet_off.max_row + 1

        buffer_id = 0
        for i in range(8, num_rows):
            # Kiểm tra tình trạng đơn
            state = sheet_off.cell(row = i, column = 8).value
            if (state != "Đã phê duyệt"):
                continue

            if (sheet_off.cell(row = i, column = 5).value != None) and ('chế độ con nhỏ' in sheet_off.cell(row = i, column = 5).value):
                continue

            id = sheet_off.cell(row = i, column = 2).value
            if id == None:
                employee = self.employees[buffer_id]
            else:
                id = int(id)
                if id not in self.employees:
                    employee = Employee(id)
                    self.employees[id] = employee
                else:
                    employee = self.employees[id]

                if buffer_id != id:
                    buffer_id = id

            dates = sheet_off.cell(row = i, column = 7).value
            dates_on_leave = dates.split(' - ')
            
            from_date_on_leave = datetime.datetime.strptime(dates_on_leave[0], "%d/%m/%Y %H:%M")
            end_date_on_leave = datetime.datetime.strptime(dates_on_leave[1], "%d/%m/%Y %H:%M")

            reason = sheet_off.cell(row = i, column = 6).value
            employee.getOnLeave(from_date_on_leave, end_date_on_leave, reason)

    def write(self):
        workbook_out = xlrd.open_workbook(self.file_output, formatting_info=True)
        r_sheet = workbook_out.sheet_by_name("Sheet1")

        wb = copy(workbook_out)
        w_sheet = wb.get_sheet(0)

        # With xlrd, indexes of col and row start with 0
        days = []
        for col in range(3, r_sheet.ncols - 1):
            day = r_sheet.cell(1, col).value
            days.append(int(day))

        for row in range(2, r_sheet.nrows - 2):
            id = r_sheet.cell(row, 1).value
            value = None

            if id not in self.employees:
                continue
            employee = self.employees[id]

            tmp = dict()
            dates = employee.Timekeeping
            for day in dates:
                date = employee.Timekeeping[day]
                if (day in self.sun) or (day in self.sat):
                    continue
                if (date.salaryCoefficient() == 0):
                    w_sheet.write(row, day + 2, "XX")
                    tmp[day] = "XX"
                if (date.salaryCoefficient() == 1):
                    w_sheet.write(row, day + 2, "X:8")
                    tmp[day] = "X:8"
                if (date.salaryCoefficient() == 2):
                    w_sheet.write(row, day + 2, "X:4")
                    tmp[day] = "X:4"

            # Viết kí hiệu vào file output
            dates_on_leave = employee.on_leave
            for date in dates_on_leave:
                month = date.month
                if (month != self.MONTH):
                    continue
                day = date.day
                if (day in self.sun) or (day in self.sat):
                    continue
                if day in tmp:
                    tmp[day] = value = tmp[day] +  ", " + dates_on_leave[date]
                    w_sheet.write(row, day + 2, value)
                else:
                    w_sheet.write(row, day + 2, dates_on_leave[date])
                    tmp[day] = dates_on_leave[date]

        wb.save(self.file_output + '[T' + str(self.MONTH) + ']' + os.path.splitext(self.file_output)[-1])

    def process(self):
        self.getExcelFiles()
        if self.file_input == None or self.file_on_leave == None or self.file_output == None:
            print("Exit")
            return
        self.read_timekeeping_machine()
        self.readFileOnLeave()
        self.write()

def main():
    timekeeper = Timekeeper()
    timekeeper.process()

if __name__ == "__main__":
    main()